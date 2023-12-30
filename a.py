import csv
import openpyxl

def matchSheng(sheng_excel,sheng):
    if sheng_excel == '于':
        return '云' == sheng
    if sheng_excel == '神':
        return '常' == sheng
    if sheng_excel == '禪':
        return '船' == sheng
    if sheng_excel == '娘':
        return '孃' == sheng
    if sheng_excel == '群':
        return '羣' == sheng
    return sheng_excel == sheng


with open('guangyun.csv', 'r', encoding='utf-8') as file:
    lines = file.readlines()[1:]

#input_yun = input("请输入韵：")
#filtered_lines = [line.strip().split(',') for line in lines if line.strip().split(',')[2] == input_yun]

processed_pairs = set()

workbook = openpyxl.load_workbook('a.xlsx')
example_sheet = workbook['example']

print('生成開始。')
for line in [line.strip().split(',') for line in lines]:
    if(len(line[3])==0): #guangyun.csv有空缺补上
        if(line[0]=='409'):
            line[3] = '匣開先上'
        elif(line[0]=='597'):
            line[3]='崇合先平'
        elif(line[0]=='646'):
            line[3]='羣合山平'
        elif(line[0]=='2021'):
            line[3]='書談上'
        elif(line[0]=='3373'):
            line[3]='透沒入'
        else:
            print(line)
            continue

    sheng = line[3][0]
    diao = line[3][-1]

    title = line[3][1:-1]
    if title not in workbook.sheetnames:
        sheet = workbook.copy_worksheet(example_sheet)
        sheet.title = title
    else:
        sheet = workbook[title]

    # 只收小韵首字
    if (title, sheng, diao) in processed_pairs:
        continue

    processed_pairs.add((title, sheng, diao))
    
    # 在excel表格中定位
    searched = False
    for row_index, row in enumerate(sheet.iter_rows(min_row=4, max_col=1, max_row=sheet.max_row)): 
        if matchSheng(row[0].value, sheng):
            for col_index, col in enumerate(sheet.iter_cols(min_col=2, max_col=5, min_row=2, max_row=2)):
                if col[0].value == diao:
                    sheet.cell(row=3, column=col_index + 2, value=line[2])
                    sheet.cell(row=row_index + 4, column=col_index + 2, value=line[6] + line[4])
                    searched = True
                    break
    if not searched:
        print('未找到：', line)

workbook.save('a.xlsx')
print('生成結束，已保存至excel。')

